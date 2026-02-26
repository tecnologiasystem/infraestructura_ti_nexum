import os
import re
import smtplib
import json
import pandas as pd
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from fastapi.responses import JSONResponse
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import base64, mimetypes
from email.header import Header
import logging
from xhtml2pdf import pisa
import traceback
from io import BytesIO
import zipfile
from dotenv import load_dotenv
import ssl
import time
from smtplib import SMTPServerDisconnected, SMTPResponseException
from email.mime.image import MIMEImage
from app.dal.Email_dal import crear_encabezado, registrar_detalle, actualizar_estado_detalle, finalizar_encabezado_si_completo, obtener_correo_usuario, actualizar_estado_encabezado, obtener_estado_encabezado, cancelar_pendientes_por_encabezado, listar_pendientes_por_encabezado, contar_enviados_hoy_por_remitente, marcar_fecha_envio_si_falta
from datetime import datetime, time as dtime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MEMBRETE_HTML_PATH = r"D:\microservicios\CorreosMasivos\app\utils\membrete_email.html"
FOOTER_HTML_PATH   = r"D:\microservicios\CorreosMasivos\app\utils\footer_email.html"

# ── Rate limiting para Microsoft 365 (sin .env) ─────────────────────────────
RATE_BLOCK_SIZE = 200          # Nº de correos que se envían antes de descansar
RATE_BLOCK_SLEEP = 900         # Descanso entre bloques (segundos) → 900 = 15 min
SLEEP_BETWEEN_MSGS = 0.5       # Micro-pausa por mensaje (0.2–0.5 si te conviene)

MAX_DAILY_PER_ACCOUNT = 600 # LIMITE DIARIO POR CUENTA
SEND_START = dtime(7, 0)  # EMPEZAR A LAS 07:00 AM RESTANTES

def _after_send_start() -> bool:
    return datetime.now().time() >= SEND_START

# ─────────────────────────────────────────────────────────────────────────────
# ENV & LOGGING
# ─────────────────────────────────────────────────────────────────────────────
load_dotenv()
logger = logging.getLogger("email_sender")
logging.basicConfig(level=logging.INFO)

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
RECIPIENT_CANDIDATES = {"correos", "correo", "email", "e-mail", "mail"}
ADJ_KEYS = {"adjuntos", "archivos"}  
ADJ_PREFIX = "adjunto"              

_MEMBRETE_HTML_CACHE = None
_FOOTER_HTML_CACHE = None

def _notify_uploader_simple(server, from_addr: str, to_addr: str, descripcion: str):
    """
    Envía un aviso simple al finalizar: muestra la descripción y la hora de finalización.
    Usa la sesión SMTP ya abierta (server) y el remitente autenticado (from_addr).
    """
    try:
        if not (server and from_addr and to_addr):
            return
        hora_fin = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        msg = MIMEMultipart("alternative")
        msg["From"] = from_addr
        msg["To"] = to_addr
        msg["Subject"] = Header("Proceso finalizado", "utf-8")

        body_txt = (
            f"Hola,\n\n"
            f"Tu carga con la descripción:\n"
            f"  \"{descripcion}\"\n\n"
            f"ha finalizado correctamente a las {hora_fin}.\n\n"
            f"Este es un mensaje automático."
        )
        body_html = (
            f"<p>Hola,</p>"
            f"<p>Tu carga con la descripción:</p>"
            f"<blockquote>{descripcion}</blockquote>"
            f"<p>ha finalizado correctamente a las <b>{hora_fin}</b>.</p>"
            f"<p><i>Este es un mensaje automático.</i></p>"
        )

        alt = MIMEMultipart("alternative")
        alt.attach(MIMEText(body_txt, "plain", "utf-8"))
        alt.attach(MIMEText(body_html, "html", "utf-8"))
        msg.attach(alt)

        server.sendmail(from_addr, [to_addr], msg.as_string())
        logging.getLogger("email_sender").info(f"✅ Notificación final enviada a {to_addr}")
    except Exception as e:
        logging.getLogger("email_sender").warning(f"❌ No se pudo notificar al finalizar ({to_addr}): {e}")

def _load_file_cached(path, cache_name):
    try:
        content = Path(path).read_text(encoding="utf-8").strip()
    except Exception as e:
        logger.warning(f"No se pudo leer {cache_name} en {path}: {e}")
        content = ""
    return content

def _load_membrete_html():
    global _MEMBRETE_HTML_CACHE
    if _MEMBRETE_HTML_CACHE is None:
        _MEMBRETE_HTML_CACHE = _load_file_cached(MEMBRETE_HTML_PATH, "membrete_html")
    return _MEMBRETE_HTML_CACHE

def _load_footer_html():
    global _FOOTER_HTML_CACHE
    if _FOOTER_HTML_CACHE is None:
        _FOOTER_HTML_CACHE = _load_file_cached(FOOTER_HTML_PATH, "footer_html")
    return _FOOTER_HTML_CACHE

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()

def _lower_no_accents(s: str) -> str:
    s = _norm(s).lower()
    tr = str.maketrans("áéíóúäëïöüàèìòùñ", "aeiouaeiouaeioun")
    return s.translate(tr)

def _pick_col_value(row_dict: dict, candidates: set[str]) -> str:
    # candidates: {"cedula","cédula","identificacion","documento", ...}
    norm_keys = { _lower_no_accents(k): k for k in row_dict.keys() }
    for c in candidates:
        nk = _lower_no_accents(c)
        if nk in norm_keys:
            val = row_dict.get(norm_keys[nk])
            return "" if val is None else str(val).strip()
    return ""

def _split_adj_list(s: str) -> List[str]:
    if not s:
        return []
    return [re.sub(r"\s+", " ", t).strip() for t in re.split(r"[;,]", str(s)) if str(t).strip()]

_IMG_SRC_RE = re.compile(r'<img[^>]+src=["\']([^"\']+)["\']', re.IGNORECASE)

def _inline_images_cid_on_msg(msg, html: str, base_dir: str) -> str:
    if not html:
        return html

    def _cid_for(name: str) -> str:
        # CID más simple y compatible
        return name.replace(".", "_").replace(" ", "_")

    out = html
    for m in list(_IMG_SRC_RE.finditer(html)):
        src = (m.group(1) or "").strip()
        if not src or src.lower().startswith(("cid:", "http://", "https://")):
            continue

        # 1) data URI (data:image/...;base64,...)
        if src.lower().startswith("data:image/"):
            try:
                header, b64 = src.split(",", 1)
                mime = header.split(";")[0].split(":", 1)[1]
                maintype, subtype = mime.split("/", 1)
                raw = base64.b64decode(b64)
                
                filename = f"membrete_inline.{subtype}"
                cid = _cid_for(filename)
                
                # Crear MIMEImage con el parámetro name (crítico para Gmail)
                img = MIMEImage(raw, _subtype=subtype, name=filename)
                img.add_header("Content-ID", f"<{cid}>")
                img.add_header("Content-Disposition", "inline", filename=filename)
                img.add_header("X-Attachment-Id", cid)
                msg.attach(img)
                out = out.replace(src, f"cid:{cid}")
                logger.info(f"✅ Imagen data URI embebida con CID: {cid}")
            except Exception as e:
                logger.warning(f"No se pudo embeber data URI del membrete: {e}")
            continue

        # 2) archivo local/relativo (logo.png, etc.)
        p = Path(src)
        if not p.is_absolute():
            p = Path(base_dir) / src
        p = p.resolve()
        
        if not p.exists():
            logger.warning(f"Imagen de membrete no encontrada: {src} -> {p}")
            continue

        try:
            mime_type, _ = mimetypes.guess_type(str(p))
            if not mime_type or not mime_type.startswith("image/"):
                mime_type = "image/png"
            
            maintype, subtype = mime_type.split("/", 1)
            
            with open(p, "rb") as f:
                img_data = f.read()
            
            filename = p.name
            cid = _cid_for(filename)
            
            # Crear MIMEImage con name (crítico para Gmail)
            img = MIMEImage(img_data, _subtype=subtype, name=filename)
            img.add_header("Content-ID", f"<{cid}>")
            img.add_header("Content-Disposition", "inline", filename=filename)
            img.add_header("X-Attachment-Id", cid)
            msg.attach(img)
            out = out.replace(src, f"cid:{cid}")
        except Exception as e:
            logger.error(f"Error embebiendo imagen {p}: {e}")
            continue
    return out


def _html_to_plain(html: str) -> str:
    if not html:
        return ""
    txt = re.sub(r"<(br|BR)\s*/?>", "\n", html)
    txt = re.sub(r"<[^>]+>", "", txt)
    txt =  re.sub(r"\r?\n\s*\r?\n+", "\n\n", txt)
    return txt.strip()


def _resolve_email_password(sender_email: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Busca en el .env un par (EMAIL_i, PASSWORD_i) cuyo EMAIL_i == sender_email.
    Devuelve (email, password) si lo encuentra; en otro caso (None, None).
    """
    if not sender_email:
        return None, None
    i = 1
    while True:
        env_email = os.getenv(f"EMAIL_{i}")
        if not env_email:
            break
        if env_email.strip().lower() == sender_email.strip().lower():
            pwd = os.getenv(f"PASSWORD_{i}")
            return env_email.strip(), (pwd.strip() if pwd else None)
        i += 1
    return None, None


def _resolve_per_doc_path_from_pattern(row: Dict, pattern: str, folder: Optional[str]) -> Optional[str]:
    """
    Usa SOLO la ruta actual: EMAIL_ATTACH_DIR.
    Si 'folder' es absoluta (UNC o disco), se usa tal cual.
    Nombre: pattern personalizado, p.ej. 'documento_{Var1}.pdf'.
    """
    if not pattern:
        return None

    # 1) Personaliza y arma nombre
    name = personalize(pattern, row).strip()
    if not name:
        return None
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    if not name.lower().endswith(".pdf"):
        name += ".pdf"

    # 2) Resuelve base
    base = (folder or "").strip()
    if base:
        if os.path.isabs(base) or base.startswith("\\\\"):
            base_dir = os.path.normpath(base)
        else:
            attach_root = os.getenv("EMAIL_ATTACH_DIR") or ""
            base_dir = os.path.normpath(os.path.join(attach_root, base))
    else:
        base_dir = os.path.normpath(os.getenv("EMAIL_ATTACH_DIR") or "")

    if not base_dir:
        logger.warning("EMAIL_ATTACH_DIR no está definido y 'folder' vacío/relativo.")
        return None

    candidate = os.path.normpath(os.path.join(base_dir, name))
    if not (os.path.isabs(folder or "") or (folder or "").startswith("\\\\")):
        base_norm = os.path.normpath(base_dir)
        if not (candidate == base_norm or candidate.startswith(base_norm + os.sep)):
            logger.warning(f"Ruta inválida (traversal): {candidate}")
            return None

    return candidate if os.path.isfile(candidate) else None

def resolve_excel_path(excel_file_name: str) -> str:
    """
    Devuelve una ruta existente para el Excel, probando:
    - ruta absoluta (si ya viene y existe)
    - EMAIL_MASIVO_DIR
    - EMAIL_ATTACH_DIR (por si quedó allí)
    Si no encuentra, retorna la mejor candidata para diagnosticar.
    """
    if not excel_file_name:
        logger.error("excel_file_name vacío")
        return ""

    # 1) Ruta absoluta
    if os.path.isabs(excel_file_name) and os.path.isfile(excel_file_name):
        return excel_file_name

    bases: List[str] = []
    base_masivo = os.getenv("EMAIL_MASIVO_DIR")
    base_attach = os.getenv("EMAIL_ATTACH_DIR")
    if base_masivo:
        bases.append(base_masivo)
    if base_attach:
        bases.append(base_attach)

    candidates: List[str] = []
    for base in bases:
        try:
            p = os.path.normpath(os.path.join(base, excel_file_name))
        except TypeError:
            continue
        candidates.append(p)
        if os.path.isfile(p):
            return p

    logger.error(
        f"⚠️ Excel NO encontrado. Probados: {candidates or ['<sin bases configuradas>']}. "
        f"ENV.EMAIL_MASIVO_DIR={base_masivo} ENV.EMAIL_ATTACH_DIR={base_attach}"
    )
    if candidates:
        return candidates[0]
    return excel_file_name

def resolve_attachments_paths(names: List[str]) -> List[str]:
    """Adjuntos comunes enviados desde el FE (mismos para todos)."""
    if not names:
        return []
    base_dir = os.getenv("EMAIL_ATTACH_DIR")
    if not base_dir:
        logger.warning("EMAIL_ATTACH_DIR no está definido; no se resolverán adjuntos comunes.")
        return []
    out = []
    for n in names:
        if not n:
            continue
        try:
            p = os.path.normpath(os.path.join(base_dir, n))
        except TypeError:
            continue
        out.append(p)
    return out

def find_recipient_column(df: pd.DataFrame) -> str:
    """
    Busca la columna de destinatarios de forma flexible (case-insensitive, sin acentos).
    """
    cols_map = {c: _lower_no_accents(c) for c in df.columns}
    for original, norm in cols_map.items():
        if norm in RECIPIENT_CANDIDATES:
            return original
    # fallback: intenta columnas que contengan 'correo' o 'email'
    for original, norm in cols_map.items():
        if "correo" in norm or "email" in norm or "mail" in norm:
            return original
    raise ValueError(
        f"No se encontró columna de destinatarios. Aceptadas: {sorted(RECIPIENT_CANDIDATES)}; "
        f"columnas disponibles: {list(df.columns)}"
    )

def personalize(text: str, row: Dict) -> str:
    """
    Reemplaza {columna} por valor usando las columnas de la fila (case-insensitive).
    """
    if not text:
        return ""
    out = text
    kv = {str(k).lower(): ("" if pd.isna(v) else str(v)) for k, v in row.items()}
    for m in re.findall(r"\{([^}]+)\}", out):
        key = m.strip().lower()
        out = out.replace("{%s}" % m, kv.get(key, ""))
    return out

def attach_files(msg: MIMEMultipart, files: List[str]) -> Tuple[List[str], List[str]]:
    ok, fail = [], []
    for fpath in files:
        try:
            if not os.path.isfile(fpath):
                fail.append(fpath)
                logger.warning(f"No existe adjunto: {fpath}")
                continue
            with open(fpath, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            filename = os.path.basename(fpath)
            part.add_header("Content-Disposition", "attachment",
                            filename=Header(filename, "utf-8").encode())
            msg.attach(part)
            ok.append(fpath)
        except Exception as e:
            fail.append(fpath)
            logger.error(f"Error adjuntando {fpath}: {e}")
    return ok, fail

def detect_is_html(body: str) -> bool:
    if not body:
        return False
    return "<" in body and ">" in body and "</" in body

def _compose_body_with_membrete(raw_body: str, row: Dict) -> tuple[str, bool]:
    membrete_html = _load_membrete_html()
    footer_html   = _load_footer_html()
    body          = personalize(raw_body, row)
    is_html       = detect_is_html(body)

    if not is_html:
        body = body.replace("\n", "<br>")
        is_html = True

    body_block = f"""
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;margin:0;padding:0;">
      <tr>
        <td style="font-family:Arial,Helvetica,sans-serif;font-size:14px;line-height:1.5;color:#222222 !important;">
          {body}
        </td>
      </tr>
    </table>
    """
    final_html = f"""{membrete_html}
{body_block}
{footer_html}"""
    return final_html, True

def resolve_row_attachments(row: Dict, base_attach_dir: str, excel_dir: Optional[str] = None) -> List[str]:
    """
    Adjuntos por persona (por fila) desde:
    - 'adjuntos' / 'archivos' con lista separada por ; o ,
    - columnas 'adjunto1', 'adjunto2', ...
    - (opcional) 'carpeta' / 'folder' subcarpeta dentro de EMAIL_ATTACH_DIR
    Soporta placeholders en nombres (ej.: informe_{cedula}.pdf).
    Orden de búsqueda por token:
      1) ruta absoluta existente
      2) EMAIL_ATTACH_DIR/carpeta/token
      3) EMAIL_ATTACH_DIR/token
      4) excel_dir/token (si se quiere buscar junto al Excel)
    """
    out: List[str] = []
    row_lc = {str(k).lower(): ("" if pd.isna(v) else v) for k, v in row.items()}
    carpeta = _norm(row_lc.get("carpeta") or row_lc.get("folder") or "")

    tokens: List[str] = []
    # 1) columnas “consolidadas”
    for k in ADJ_KEYS:
        tokens += _split_adj_list(row_lc.get(k, ""))

    # 2) adjunto1..n
    for k, v in row.items():
        k_l = str(k).lower()
        if k_l.startswith(ADJ_PREFIX):
            tokens += _split_adj_list(v)

    # Render placeholders
    tokens = [personalize(t, row) for t in tokens]

    for tok in tokens:
        if not tok:
            continue
        tok = tok.replace("/", "\\").strip()

        # a) absoluta
        if os.path.isabs(tok) and os.path.isfile(tok):
            out.append(os.path.normpath(tok))
            continue

        # b) EMAIL_ATTACH_DIR/carpeta/tok
        if base_attach_dir and carpeta:
            p = os.path.normpath(os.path.join(base_attach_dir, carpeta, tok))
            if os.path.isfile(p):
                out.append(p); continue

        # c) EMAIL_ATTACH_DIR/tok
        if base_attach_dir:
            p = os.path.normpath(os.path.join(base_attach_dir, tok))
            if os.path.isfile(p):
                out.append(p); continue

        # d) junto al Excel
        if excel_dir:
            p = os.path.normpath(os.path.join(excel_dir, tok))
            if os.path.isfile(p):
                out.append(p); continue

        logger.warning(f"No se encontró adjunto para fila: '{tok}' (carpeta='{carpeta}')")

    # Dedup
    seen, dedup = set(), []
    for p in out:
        if p not in seen:
            dedup.append(p); seen.add(p)
    return dedup

# ——— HELPERS SMTP ———
def _mk_smtp(smtp_server, smtp_port, email_user, email_password, timeout=120):
    """
    Crea una sesión SMTP robusta: EHLO → STARTTLS → EHLO → LOGIN.
    Usa timeout mayor para redes lentas o adjuntos grandes.
    """
    srv = smtplib.SMTP(smtp_server, smtp_port, timeout=timeout)
    srv.ehlo()
    ctx = ssl.create_default_context()
    srv.starttls(context=ctx)
    srv.ehlo()
    print("====== SMTP LOGIN DEBUG ======")
    print("SMTP host     :", smtp_server)
    print("SMTP port     :", smtp_port)
    print("SMTP user     :", email_user)
    print("SMTP password :", email_password)
    print("================================")

    srv.login(email_user, email_password)
    return srv

def _send_final_notification(email_user, email_password, to_addr, descripcion):
    try:
        smtp_server = os.getenv("SERVER", "smtp.office365.com")
        smtp_port = int(os.getenv("PORT", "587"))

        server = _mk_smtp(smtp_server, smtp_port, email_user, email_password, timeout=60)

        _notify_uploader_simple(
            server=server,
            from_addr=email_user,
            to_addr=to_addr,
            descripcion=descripcion
        )

        _smtp_safe_quit(server)
    except Exception as e:
        logger.warning(f"❌ No se pudo notificar al finalizar ({to_addr}): {e}")


def _smtp_safe_quit(server):
    try:
        server.quit()
    except Exception:
        pass

def _smtp_noop(server, logger=None):
    try:
        server.noop()
        return True
    except Exception as e:
        if logger:
            logger.warning(f"NOOP falló: {e}")
        return False

def _send_with_retry(server, mk_server, msg, recipient, envelope_from, logger=None, per_recipient_tries=2):
    tries = 0
    while True:
        try:
            result = server.sendmail(
                envelope_from,
                [recipient],
                msg.as_bytes()   
            )   

            if result:
                if logger:
                    logger.error(f"SMTP RCPT error for {recipient}: {result}")
                return server, False, f"RCPT refused: {result}"
            return server, True, None

        except SMTPServerDisconnected as e:
            tries += 1
            if logger:
                logger.warning(f"Conexión SMTP cerrada al enviar a {recipient}. Reintento {tries}/{per_recipient_tries}…")
            if tries >= per_recipient_tries:
                return server, False, f"SMTP desconectado: {e}"
            _smtp_safe_quit(server)
            server = mk_server()

        except SMTPResponseException as e:
            # 👇 clave: si el server queda "descuadrado", resetea o reconecta
            code = getattr(e, "smtp_code", None)
            tries += 1

            if logger:
                logger.warning(f"SMTPResponseException {code} para {recipient}: {e}")

            # 503 5.5.1 = secuencia mala -> RSET o reconectar
            if code == 503:
                try:
                    server.rset()  # resetea estado de transacción
                except Exception:
                    _smtp_safe_quit(server)
                    server = mk_server()

                if tries < per_recipient_tries:
                    continue
                return server, False, f"{code} Bad sequence: {e}"

            # Para otros códigos: no siempre conviene reconectar
            return server, False, str(e)

        except Exception as e:
            # si es error raro, mejor reconectar una vez
            tries += 1
            if tries < per_recipient_tries:
                try:
                    _smtp_safe_quit(server)
                    server = mk_server()
                    continue
                except Exception:
                    pass
            return server, False, str(e)
        
# RESPETA EL FORMATO MONEDA
def leer_excel_respetando_formato(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows = []
    headers = [c.value for c in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=False):
        row_dict = {}
        for idx, cell in enumerate(row):
            col_name = headers[idx]
            value = cell.value

            if isinstance(value, (int, float)):
                number_format = (cell.number_format or "").lower()

                # FORMATO SOLO SI ES MONEDA
                if "$" in number_format or "[$" in number_format:
                    # Redondear
                    entero = int(round(value))
                    txt = f"{entero:,.0f}"
                    txt = txt.replace(",", ".")
                    value = f"$ {txt}"
                else:
                    # Si NO es moneda lo dejo TAL CUAL
                    value = str(value).replace(".0", "")

            row_dict[col_name] = value

        rows.append(row_dict)

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# CORE
# ─────────────────────────────────────────────────────────────────────────────
def enviar_correos_masivos(data: dict):
    # Payload
    subject_tmpl = _norm(data.get("subject"))
    body_tmpl = data.get("body") or ""
    excel_name = _norm(data.get("excelFileName"))
    attachment_names = data.get("attachments") or []
    attachments_mode = (data.get("attachmentsMode") or "both").lower().strip()
    is_html = bool(data.get("isHtml", detect_is_html(body_tmpl)))
    per_doc_pattern = _norm(data.get("perDocPattern"))
    per_doc_folder  = _norm(data.get("perDocFolder"))

    if not subject_tmpl or not body_tmpl:
        return JSONResponse({"error": "Faltan subject o body"}, status_code=400)
    if not excel_name:
        return JSONResponse({"error": "Falta excelFileName"}, status_code=400)

    # Excel
    excel_path = resolve_excel_path(excel_name)
    if not os.path.isfile(excel_path):
        detail = {
            "resolved_path": excel_path,
            "exists": os.path.isfile(excel_path),
            "EMAIL_MASIVO_DIR": os.getenv("EMAIL_MASIVO_DIR"),
            "EMAIL_ATTACH_DIR": os.getenv("EMAIL_ATTACH_DIR"),
            "excelFileName": excel_name,
        }
        logger.error(f"❌ Excel no encontrado: {detail}")
        return JSONResponse({"error": "No existe el Excel", "detail": detail}, status_code=404)

    try:
        rows = leer_excel_respetando_formato(excel_path)
        df = pd.DataFrame(rows)

        logger.info(f"✅ Excel leído: filas={len(df)} columnas={list(df.columns)}")
    except Exception as e:
        tb = traceback.format_exc()
        hint = "Sugerencia: pip install openpyxl" if "openpyxl" in str(e).lower() else ""
        logger.error(f"Error leyendo Excel: {e}\n{tb}")
        return JSONResponse({"error": f"Error leyendo Excel: {e}", "hint": hint}, status_code=500)

    if df.empty:
        return JSONResponse({"error": "El Excel no contiene filas"}, status_code=404)
    
    sender_email = _norm(data.get("senderEmail"))
    total_registros = len(df)
    id_usuario = (data.get("userId") or data.get("idUsuario") or None)
    descripcion = f"{sender_email or 'remitente_desconocido'} | {os.path.basename(excel_path)}"
    id_encabezado = crear_encabezado(descripcion, id_usuario, total_registros, sender_email)
    actualizar_estado_encabezado(id_encabezado, "EN_PROCESO")

    try:
        recipient_col = find_recipient_column(df)
        logger.info(f"📬 Columna destinatarios: '{recipient_col}'")
    except ValueError as ve:
        return JSONResponse({"error": str(ve), "columns": list(df.columns)}, status_code=400)
    
    # ✅ FASE 1: PRECARGAR TODO A BD (PENDIENTE) ANTES DE ENVIAR
    for idx, row in df.iterrows():
        row_dict = {str(k): ("" if pd.isna(v) else v) for k, v in row.to_dict().items()}
        recipient = _norm(row_dict.get(recipient_col))

        if recipient and (";" in recipient or "," in recipient):
            recipient = _norm(re.split(r"[;,]", recipient)[0])

        if not recipient:
            continue

        # Renderiza asunto y cuerpo igual que lo haces hoy
        subject = personalize(subject_tmpl, row_dict)
        final_body, _ = _compose_body_with_membrete(body_tmpl, row_dict)

        # Serializa adjuntos (si quieres guardar los que aplican por fila)
        # Si no quieres calcular adjuntos aquí, puedes guardar None
        adjuntos_str = None

        cedula = _pick_col_value(row_dict, {"cedula", "cédula", "identificacion", "documento", "dni", "cc"})
        nombre_cliente = _pick_col_value(row_dict, {"nombre", "nombre_cliente", "cliente", "nombre del cliente"})
        campana = _pick_col_value(row_dict, {"campaña", "campana", "campaign"})

        # Inserta detalle PENDIENTE (solo BD, sin enviar)
        registrar_detalle(
            id_encabezado,
            recipient,
            subject,
            final_body,
            adjuntos_str,
            cedula,
            nombre_cliente,
            campana
        )

    logger.info(f"📥 Precarga completada en BD para encabezado={id_encabezado}")

    # 1) intenta emparejar EMAIL_i/PASSWORD_i por senderEmail
    email_user, email_password = _resolve_email_password(sender_email)

    # 2) fallback legacy si no encontró par
    if not email_user or not email_password:
        email_user = email_user or os.getenv("REMITENTENPL")
        email_password = email_password or os.getenv("PASSWORDNPL")

    smtp_server = os.getenv("SERVER", "smtp.office365.com")
    smtp_port = int(os.getenv("PORT", "587"))

    # Validaciones claras
    if not sender_email:
        return JSONResponse({"error": "Debe enviar senderEmail"}, status_code=400)
    if not email_user:
        return JSONResponse({"error": f"senderEmail '{sender_email}' no está configurado (EMAIL_i) y no hay fallback REMITENTENPL"}, status_code=500)
    if not email_password:
        return JSONResponse({"error": f"Falta PASSWORD_i para '{email_user}' o PASSWORDNPL (fallback)"}, status_code=500)
    
    if not _after_send_start():
        return JSONResponse({
            "message": "Aún no son las 07:00. Queda en cola y se retomará automáticamente desde las 07:00.",
            "idEncabezado": id_encabezado
        }, status_code=200)
    
    # ── Parámetros de lote/keepalive
    BATCH_SIZE = int(os.getenv("EMAIL_SMTP_BATCH_SIZE", "75"))      # 50–100 recomendado
    NOOP_EVERY = int(os.getenv("EMAIL_SMTP_NOOP_EVERY", "15"))      # ping cada 10–20
    SLEEP_BETWEEN_BATCH = float(os.getenv("EMAIL_SMTP_SLEEP", "1")) # 1–2 s entre lotes

    try:
        npl_user = os.getenv("REMITENTENPL") or ""
        if email_user and npl_user and email_user.lower() == npl_user.lower():
            t = _mk_smtp(smtp_server, smtp_port, npl_user, os.getenv("PASSWORDNPL"))
            _smtp_safe_quit(t)
    except Exception as e:
        logger.error(f"Test SMTP NPL FAIL: {e}")

    # ── Conexión inicial SMTP
    try:
        server = _mk_smtp(smtp_server, smtp_port, email_user, email_password, timeout=120)
        server.set_debuglevel(1)
        logger.info("Conexión SMTP OK.")
    except Exception as e:
        logger.error(f"Error SMTP: {e}")
        return JSONResponse({"error": f"Error SMTP: {e}"}, status_code=500)

    # Factory para reconectar cuando el servidor corte
    def _mk():
        return _mk_smtp(smtp_server, smtp_port, email_user, email_password, timeout=120)

    # ── Adjuntos comunes
    attachment_paths = resolve_attachments_paths(attachment_names)
    common_ok, common_fail = [], []
    if attachments_mode in ("common", "both") and attachment_paths:
        dummy = MIMEMultipart()
        common_ok, common_fail = attach_files(dummy, attachment_paths)
        if common_fail:
            logger.warning(f"Adjuntos comunes no encontrados: {common_fail}")
    else:
        common_ok = []

    base_attach_dir = os.getenv("EMAIL_ATTACH_DIR") or ""
    excel_dir = os.path.dirname(excel_path) if excel_path else None

    sent, errors = 0, []

    try:
        block_sent = 0

        enviados_hoy = contar_enviados_hoy_por_remitente(email_user)

        while True:
            # 🔴 Si pausaron el encabezado, salimos para liberar worker (como ya lo hiciste)
            estado = obtener_estado_encabezado(id_encabezado)
            if estado == "PAUSADO":
                print(f"⏸️ Envío PAUSADO para encabezado {id_encabezado}. Liberando worker...")
                try:
                    _smtp_safe_quit(server)
                except Exception:
                    pass
                return JSONResponse({"message": "Proceso pausado por el usuario", "idEncabezado": id_encabezado}, status_code=200)

            if estado in ("CANCELADO", "CANCELADO_POR_USUARIO"):
                logger.warning(f"🛑 Envío CANCELADO para encabezado {id_encabezado}")
                try:
                    cancelar_pendientes_por_encabezado(id_encabezado)
                except Exception:
                    pass
                return JSONResponse({"message": "Proceso cancelado por el usuario", "idEncabezado": id_encabezado}, status_code=200)

            # Trae lista de pendientes (puedes optimizar a TOP 1 después)
            pendientes = listar_pendientes_por_encabezado(id_encabezado)
            if not pendientes:
                break

            for p in pendientes:
                id_detalle = p["idDetalle"]
                recipient = _norm(p["email_destinatario"])
                subject   = p["asunto"] or ""
                final_body = p["cuerpo"] or ""
                adjuntos_str = p.get("adjuntos")

                # LÍMITE DIARIO (ANTES DE ENVIAR ESTE CORREO)
                if enviados_hoy >= MAX_DAILY_PER_ACCOUNT:
                    logger.info(f"⛔ Límite diario alcanzado ({MAX_DAILY_PER_ACCOUNT}) para {email_user}. Continúa mañana desde las 07:00.")
                    try:
                        _smtp_safe_quit(server)
                    except Exception:
                        pass
                    # NO PAUSAR. Queda EN_PROCESO con PENDIENTES.
                    return JSONResponse({
                        "message": "Límite diario alcanzado. Los pendientes se enviarán automáticamente mañana desde las 07:00.",
                        "idEncabezado": id_encabezado,
                        "sent_today": enviados_hoy,
                        "limit": MAX_DAILY_PER_ACCOUNT
                    }, status_code=200)

                try:
                    msg = MIMEMultipart("related")
                    msg["From"] = email_user
                    if sender_email and sender_email.lower() != email_user.lower():
                        msg["Reply-To"] = sender_email
                    msg["To"] = recipient
                    msg["Subject"] = Header(subject, "utf-8")

                    alt = MIMEMultipart("alternative")
                    msg.attach(alt)

                    # Como ya guardaste HTML final en BD, lo tratamos como HTML
                    plain_part = _html_to_plain(final_body)
                    alt.attach(MIMEText(plain_part, "plain", "utf-8"))

                    base_dir = str(Path(MEMBRETE_HTML_PATH).parent)
                    attach_dir = (os.getenv("EMAIL_ATTACH_DIR") or "").strip()
                    if attach_dir:
                        def _inline_with_multi_base(msg, html, bases):
                            out = html
                            for b in bases:
                                out = _inline_images_cid_on_msg(msg, out, b)
                            return out
                        html_with_cid = _inline_with_multi_base(msg, final_body, [base_dir, attach_dir])
                    else:
                        html_with_cid = _inline_images_cid_on_msg(msg, final_body, base_dir)
                    alt.attach(MIMEText(html_with_cid, "html", "utf-8"))

                    # Si guardaste adjuntos_str (JSON), aquí puedes re-adjuntar.
                    # Si no, no adjuntas nada.
                    # (para no inventar: lo dejamos sin adjuntos por ahora)

                    # justo antes de _send_with_retry(...)
                    if attachments_mode in ("common", "both") and common_ok:
                        attach_files(msg, common_ok)

                    server, ok, err = _send_with_retry(server, _mk, msg, recipient, email_user, logger=logger, per_recipient_tries=2)
                    if not ok:
                        actualizar_estado_detalle(id_detalle, "ERROR", str(err))
                        errors.append({"idDetalle": id_detalle, "to": recipient, "error": str(err)})
                    else:
                        actualizar_estado_detalle(id_detalle, "ENVIADO", None)
                        marcar_fecha_envio_si_falta(id_detalle)   # para que cuente en el día
                        enviados_hoy += 1                         # suma SOLO si fue efectivo
                        sent += 1
                        block_sent += 1

                        if SLEEP_BETWEEN_MSGS > 0:
                            time.sleep(SLEEP_BETWEEN_MSGS)

                        if RATE_BLOCK_SIZE > 0 and block_sent >= RATE_BLOCK_SIZE:
                            logger.info(f"⏸️ Pausa anti-spam: {RATE_BLOCK_SLEEP}s tras {RATE_BLOCK_SIZE} envíos…")
                            _smtp_safe_quit(server)
                            time.sleep(RATE_BLOCK_SLEEP)
                            server = _mk()
                            block_sent = 0

                        if NOOP_EVERY > 0 and (sent % NOOP_EVERY == 0):
                            _smtp_noop(server, logger=logger)

                        if BATCH_SIZE > 0 and (sent % BATCH_SIZE == 0):
                            logger.info(f"Lote {BATCH_SIZE} completado. Reconectando…")
                            _smtp_safe_quit(server)
                            time.sleep(SLEEP_BETWEEN_BATCH)
                            server = _mk()

                except Exception as e:
                    actualizar_estado_detalle(id_detalle, "ERROR", str(e))
                    errors.append({"idDetalle": id_detalle, "to": recipient, "error": str(e)})

        actualizar_estado_encabezado(id_encabezado, "FINALIZADO")

        try:
            correo_usuario = obtener_correo_usuario(id_usuario)
            if correo_usuario:
                _send_final_notification(
                    email_user=email_user,
                    email_password=email_password,
                    to_addr=correo_usuario,
                    descripcion=descripcion
                )
        except Exception as e:
            logger.warning(f"No se pudo enviar correo de finalización: {e}")


        # Termina normal
        return JSONResponse(
            {"message": "Proceso finalizado", "idEncabezado": id_encabezado, "sent": sent, "failed": len(errors), "errors": errors[:200]},
            status_code=200 if sent > 0 else 500
        )

    finally:
        _smtp_safe_quit(server)



def _quill_classes_to_inline(html: str) -> str:
    # Traduce clases típicas de Quill a estilos inline (seguro, rápido)
    html = re.sub(r'class="([^"]*?)\bql-align-right\b([^"]*?)"', r'style="text-align:right"', html)
    html = re.sub(r'class="([^"]*?)\bql-align-center\b([^"]*?)"', r'style="text-align:center"', html)
    html = re.sub(r'class="([^"]*?)\bql-align-justify\b([^"]*?)"', r'style="text-align:justify"', html)
    # indents básicos (opcional)
    html = re.sub(r'class="([^"]*?)\bql-indent-1\b([^"]*?)"', r'style="margin-left: 2em"', html)
    html = re.sub(r'class="([^"]*?)\bql-indent-2\b([^"]*?)"', r'style="margin-left: 4em"', html)
    return html

def _wrap_html(html: str) -> str:
    # Normaliza clases de Quill a inline para que el motor PDF las respete
    html = _quill_classes_to_inline(html)
    # CSS de apoyo por si quedó alguna clase
    css = """
      @page { size: A4; margin: 2.5cm; }
      body { font-family: DejaVu Sans, Arial, Helvetica, sans-serif; font-size: 12pt; color: #000; }
      p { margin: 0 0 10px 0; }
      ul, ol { margin: 0 0 10px 1.2em; }
      .ql-align-right { text-align: right; }
      .ql-align-center { text-align: center; }
      .ql-align-justify { text-align: justify; }
      .ql-indent-1 { margin-left: 2em; }
      .ql-indent-2 { margin-left: 4em; }
    """
    return f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8"/>
  <style>{css}</style>
</head>
<body>
{html}
</body>
</html>"""

def _html_to_pdf_bytes(html: str) -> bytes:
    html = _wrap_html(html)          
    out = BytesIO()
    result = pisa.CreatePDF(html, dest=out)  
    if getattr(result, "err", 1) != 0:
        raise RuntimeError("xhtml2pdf no pudo renderizar el HTML")
    return out.getvalue()

def generar_documentos_personalizados_zip(data: dict):
    """
    data = {
      "excelFileName": "plantilla.xlsx",
      "templateHtml": "<p>Hola {Var1}</p>",
      "fileNameTemplate": "doc_{Var1}.pdf",
      "output": "pdf"
    }
    Devuelve (zip_bytes, error_str|None)
    """
    try:
        excel_name = _norm(data.get("excelFileName"))
        template_html = data.get("templateHtml") or ""
        file_name_template = _norm(data.get("fileNameTemplate") or "documento_{Var1}.pdf")

        if not excel_name or not template_html:
            return None, "Faltan excelFileName o templateHtml"

        excel_path = resolve_excel_path(excel_name)
        if not os.path.isfile(excel_path):
            return None, f"No existe Excel en: {excel_path}"

        df = pd.read_excel(excel_path, dtype=str, keep_default_na=False)

        if df.empty:
            return None, "El Excel no contiene filas"

        mem_zip = BytesIO()
        with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for _, row in df.iterrows():
                row_dict = {str(k): ("" if pd.isna(v) else v) for k, v in row.to_dict().items()}

                # Renderiza plantilla HTML y nombre por fila
                html = personalize(template_html, row_dict)
                fname = personalize(file_name_template, row_dict).strip() or "documento.pdf"
                if not fname.lower().endswith(".pdf"):
                    fname += ".pdf"

                pdf_bytes = _html_to_pdf_bytes(html)
                zf.writestr(fname, pdf_bytes)

        mem_zip.seek(0)
        return mem_zip.getvalue(), None

    except Exception as e:
        tb = traceback.format_exc()
        logger.error(f"Error generando documentos: {e}\n{tb}")
        return None, str(e)

def reanudar_envio_por_encabezado(id_encabezado: int, sender_email: str, user_id: int | None):
    # Pasa a EN_PROCESO
    actualizar_estado_encabezado(id_encabezado, "EN_PROCESO")

    pendientes = listar_pendientes_por_encabezado(id_encabezado)
    if not pendientes:
        actualizar_estado_encabezado(id_encabezado, "FINALIZADO")
        return {"message": "No hay pendientes", "idEncabezado": id_encabezado}

    # 1) credenciales según senderEmail (igual que en enviar_correos_masivos)
    email_user, email_password = _resolve_email_password(sender_email)

    if not email_user or not email_password:
        email_user = email_user or os.getenv("REMITENTENPL")
        email_password = email_password or os.getenv("PASSWORDNPL")

    smtp_server = os.getenv("SERVER", "smtp.office365.com")
    smtp_port = int(os.getenv("PORT", "587"))

    if not sender_email:
        return JSONResponse({"error": "Debe enviar senderEmail"}, status_code=400)
    if not email_user:
        return JSONResponse(
            {"error": f"senderEmail '{sender_email}' no está configurado (EMAIL_i) y no hay fallback REMITENTENPL"},
            status_code=500
        )
    if not email_password:
        return JSONResponse({"error": f"Falta PASSWORD_i para '{email_user}' o PASSWORDNPL (fallback)"}, status_code=500)

    # Parámetros
    NOOP_EVERY = int(os.getenv("EMAIL_SMTP_NOOP_EVERY", "15"))
    SLEEP_BETWEEN_MSGS = float(os.getenv("EMAIL_SMTP_SLEEP_PER_MSG", "0.2"))

    # ✅ No enviar antes de las 07:00
    if not _after_send_start():
        return JSONResponse({
            "message": "Aún no son las 07:00. Se retomará automáticamente desde las 07:00.",
            "idEncabezado": id_encabezado
        }, status_code=200)


    # ✅ Conteo efectivo del día (solo ENVIADO cuenta)
    enviados_hoy = contar_enviados_hoy_por_remitente(email_user)

    # Conecta SMTP
    try:
        server = _mk_smtp(smtp_server, smtp_port, email_user, email_password, timeout=120)
        server.set_debuglevel(1)
        logger.info(f"Conexión SMTP OK (resume). encabezado={id_encabezado}")
    except Exception as e:
        logger.error(f"Error SMTP (resume): {e}")
        return JSONResponse({"error": f"Error SMTP: {e}"}, status_code=500)

    base_dir = str(Path(MEMBRETE_HTML_PATH).parent)
    attach_dir = (os.getenv("EMAIL_ATTACH_DIR") or "").strip()

    sent = 0
    errors = 0
    try:
        while True:
            estado = obtener_estado_encabezado(id_encabezado)

            if estado == "PAUSADO":
                logger.info(f"⏸️ Resume detenido: encabezado {id_encabezado} quedó PAUSADO.")
                try:
                    _smtp_safe_quit(server)
                except Exception:
                    pass
                return {"message": "Proceso pausado por el usuario", "idEncabezado": id_encabezado}

            if estado in ("CANCELADO", "CANCELADO_POR_USUARIO"):
                logger.warning(f"🛑 Resume detenido: encabezado {id_encabezado} CANCELADO.")
                try:
                    cancelar_pendientes_por_encabezado(id_encabezado)
                except Exception:
                    pass
                return {"message": "Proceso cancelado por el usuario", "idEncabezado": id_encabezado}

            pendientes = listar_pendientes_por_encabezado(id_encabezado)
            if not pendientes:
                break

            for p in pendientes:
                id_detalle = p["idDetalle"]
                recipient = _norm(p.get("email_destinatario") or "")
                subject = p.get("asunto") or ""
                final_body = p.get("cuerpo") or ""

                if not recipient:
                    try:
                        actualizar_estado_detalle(id_detalle, "ERROR", "Sin email_destinatario")
                    except Exception:
                        pass
                    errors += 1
                    continue

                # ✅ LÍMITE DIARIO (ANTES DE ENVIAR ESTE CORREO)
                if enviados_hoy >= MAX_DAILY_PER_ACCOUNT:
                    logger.info(f"⛔ Límite diario alcanzado ({MAX_DAILY_PER_ACCOUNT}) para {email_user}. Continúa mañana desde las 07:00.")
                    try:
                        _smtp_safe_quit(server)
                    except Exception:
                        pass

                    return {
                        "message": "Límite diario alcanzado. Los pendientes continúan automáticamente mañana desde las 07:00.",
                        "idEncabezado": id_encabezado,
                        "sent_today": enviados_hoy,
                        "limit": MAX_DAILY_PER_ACCOUNT
                    }


                try:
                    msg = MIMEMultipart("related")
                    msg["From"] = email_user
                    if sender_email and sender_email.lower() != email_user.lower():
                        msg["Reply-To"] = sender_email
                    msg["To"] = recipient
                    msg["Subject"] = Header(subject, "utf-8")

                    alt = MIMEMultipart("alternative")
                    msg.attach(alt)

                    plain_part = _html_to_plain(final_body)
                    alt.attach(MIMEText(plain_part, "plain", "utf-8"))

                    # Inline imágenes (membrete / adjuntos dir)
                    if attach_dir:
                        def _inline_with_multi_base(m, html, bases):
                            out = html
                            for b in bases:
                                out = _inline_images_cid_on_msg(m, out, b)
                            return out
                        html_with_cid = _inline_with_multi_base(msg, final_body, [base_dir, attach_dir])
                    else:
                        html_with_cid = _inline_images_cid_on_msg(msg, final_body, base_dir)

                    alt.attach(MIMEText(html_with_cid, "html", "utf-8"))

                    # Adjuntos por registro (si los guardaste en BD)
                    adjuntos_str = p.get("adjuntos")
                    if adjuntos_str:
                        try:
                            # soporta JSON lista o string simple
                            files = json.loads(adjuntos_str) if isinstance(adjuntos_str, str) and adjuntos_str.strip().startswith("[") else adjuntos_str
                            if isinstance(files, str):
                                files = [x.strip() for x in re.split(r"[;,]", files) if x.strip()]
                            if isinstance(files, list):
                                # si vienen nombres, resuelve rutas con tu helper
                                for fpath in resolve_attachments_paths(files):
                                    attach_files(msg, [fpath])
                        except Exception:
                            # si falla parseo, no revientes el envío
                            pass

                    server.sendmail(email_user, [recipient], msg.as_string())

                    actualizar_estado_detalle(id_detalle, "ENVIADO", None)
                    marcar_fecha_envio_si_falta(id_detalle)
                    enviados_hoy += 1
                    sent += 1


                    if NOOP_EVERY and (sent % NOOP_EVERY == 0):
                        try:
                            server.noop()
                        except Exception:
                            try:
                                _smtp_safe_quit(server)
                            except Exception:
                                pass
                            server = _mk_smtp(smtp_server, smtp_port, email_user, email_password, timeout=120)

                    if SLEEP_BETWEEN_MSGS:
                        time.sleep(SLEEP_BETWEEN_MSGS)

                except Exception as e:
                    errors += 1
                    try:
                        actualizar_estado_detalle(id_detalle, "ERROR", str(e))
                    except Exception:
                        pass

        # Finaliza encabezado si ya no hay pendientes
        try:
            finalizar_encabezado_si_completo(id_encabezado)
        except Exception:
            pass

        actualizar_estado_encabezado(id_encabezado, "FINALIZADO")

        try:
            correo_usuario = obtener_correo_usuario(user_id)
            if correo_usuario:
                _notify_uploader_simple(
                    server=server,
                    from_addr=email_user,
                    to_addr=correo_usuario,
                    descripcion=f"Reanudación envío #{id_encabezado}"
                )
        except Exception as e:
            logger.warning(f"No se pudo enviar correo de finalización (resume): {e}")



        return {"message": "Reanudación completada", "idEncabezado": id_encabezado, "sent": sent, "errors": errors}

    finally:
        try:
            _smtp_safe_quit(server)
        except Exception:
            pass
