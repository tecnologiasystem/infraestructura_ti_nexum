from io import BytesIO
from typing import Dict, Optional, Any
import pandas as pd
import warnings
from openpyxl import load_workbook

from app.dal.acuerdoPago_dal import insertar_acuerdos_lote, obtener_dni, pop_dni, pop_acuerdo

def _s(v) -> str:
    return "" if v is None or pd.isna(v) else str(v).strip()

def _validate_excel_file(file_bytes: bytes) -> bool:
    """
    Valida que el archivo Excel sea válido y tenga contenido
    """
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        
        if len(workbook.worksheets) == 0:
            raise ValueError("El archivo Excel no contiene hojas de trabajo")
        
        # Verificar que la primera hoja tenga datos
        first_sheet = workbook.worksheets[0]
        if first_sheet.max_row <= 1:
            raise ValueError("La hoja de trabajo no contiene datos (solo encabezados o está vacía)")
        
        workbook.close()
        return True
        
    except Exception as e:
        print(f"Error validando Excel con openpyxl: {e}")
        return False

def insertar_excel_acuerdos_bytes(file_bytes: bytes) -> Dict[str, int]:
    """
    Lee el Excel (primera hoja) y llama el SP por cada fila.
    Versión robusta con múltiples engines y validaciones.
    """
    try:
        if not _validate_excel_file(file_bytes):
            raise ValueError("El archivo Excel está corrupto o no tiene el formato correcto")
        
        
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            
            df = None
            engines_to_try = ['openpyxl', 'xlrd']
            
            for engine in engines_to_try:
                try:
                    print(f"Intentando leer Excel con engine: {engine}")
                    
                    if engine == 'openpyxl':
                        df = pd.read_excel(
                            BytesIO(file_bytes), 
                            engine='openpyxl',
                            dtype=str,
                            na_values=['', 'NULL', 'null', 'None'],
                            keep_default_na=False
                        )
                    elif engine == 'xlrd':
                        df = pd.read_excel(
                            BytesIO(file_bytes),
                            engine='xlrd', 
                            dtype=str,
                            na_values=['', 'NULL', 'null', 'None'],
                            keep_default_na=False
                        )
                    
                    if df is not None and not df.empty:
                        print(f"✅ Excel leído exitosamente con {engine}")
                        break
                        
                except Exception as e:
                    print(f"❌ Error con engine {engine}: {e}")
                    continue
            
            # Si pandas falló, intentar con openpyxl directo
            if df is None or df.empty:
                print("Intentando lectura manual con openpyxl...")
                try:
                    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
                    worksheet = workbook.active
                    data = []
                    headers = None
                    
                    for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
                        if row_idx == 0:
                            headers = [str(cell) if cell is not None else f"Col_{i}" for i, cell in enumerate(row)]
                        else:
                            if any(cell is not None for cell in row):  # Solo filas con datos
                                data.append([str(cell) if cell is not None else "" for cell in row])
                    
                    if headers and data:
                        df = pd.DataFrame(data, columns=headers)
                        print("✅ Excel leído manualmente con openpyxl")
                    
                    workbook.close()
                    
                except Exception as e:
                    print(f"❌ Error con lectura manual: {e}")
        
        if df is None or df.empty:
            raise ValueError("No se pudo leer el archivo Excel con ningún método disponible")
        
        df = df.fillna("")
        
        if len(df) > 0:
            df.head(2).to_string()
        
        # Mapeo de nombres de columnas Excel -> Código
        column_mapping = {
            "Número DNI": "NumeroDNI",
            
            "Idenficacion Asesor": "IdentificacionAsesor",
            "Identificacion Asesor": "IdentificacionAsesor", 
            
            "Codigo Estado": "CodigoEstado",
            "CodigoEstado": "CodigoEstado",
            
            "Codigo Canal": "CodigoCanal",
            "CodigoCanal": "CodigoCanal",
            
            "Codigo Gestion": "CodigoGestion", 
            "CodigoGestion": "CodigoGestion",
            
            "Telefono Contactato": "TelefonoContacto",
            "Telefono Contacto": "TelefonoContacto",
            "TelefonoContacto": "TelefonoContacto",
            
            "Dirección": "Direccion",
            "Direccion": "Direccion",
            
            "Fecha promesa": "FechaPromesa",
            "Fecha Promesa": "FechaPromesa",
            "FechaPromesa": "FechaPromesa",
            
            "Valor promesa": "ValorPromesa",
            "Valor Promesa": "ValorPromesa", 
            "ValorPromesa": "ValorPromesa",
            
            "Nro de Obligacion": "NumeroObligacion",
            "Numero de Obligacion": "NumeroObligacion",
            "NumeroObligacion": "NumeroObligacion",
            
            "Fecha y hora gestión": "FechaHoraGestion",
            "Fecha y Hora Gestion": "FechaHoraGestion",
            "FechaHoraGestion": "FechaHoraGestion",
            
            "Observaciones": "Observaciones"
        }
        
        columns_mapped = {}
        for excel_col in df.columns:
            mapped_col = column_mapping.get(excel_col.strip())
            if mapped_col:
                columns_mapped[excel_col] = mapped_col
            else:
                excel_clean = excel_col.lower().replace(" ", "").replace("ó", "o").replace("é", "e")
                for map_key, map_value in column_mapping.items():
                    key_clean = map_key.lower().replace(" ", "").replace("ó", "o").replace("é", "e")
                    if excel_clean == key_clean:
                        columns_mapped[excel_col] = map_value
                        break
        
        
        # Renombrar columnas encontradas
        df = df.rename(columns=columns_mapped)
        
        # Verificar columnas requeridas
        required_cols = ["NumeroDNI", "IdentificacionAsesor", "CodigoEstado", "CodigoCanal", 
                        "CodigoGestion", "TelefonoContacto", "FechaPromesa", "ValorPromesa", 
                        "NumeroObligacion", "FechaHoraGestion", "Observaciones"]
        
        missing_cols = []
        for col in required_cols:
            if col not in df.columns:
                df[col] = ""
                missing_cols.append(col)
        
        if "Direccion" not in df.columns:
            df["Direccion"] = ""
            
        dni_list = []
        if "NumeroDNI" in df.columns:
            dni_series = df["NumeroDNI"].astype(str).map(_s)
            dni_list = [d for d in dni_series.unique().tolist() if d]  # únicos y no vacíos
            
        param_rows = []
        for idx, row in df.iterrows():
            try:
                params = (
                    1,                                          
                    None,                                       
                    _s(row["NumeroDNI"]),
                    _s(row["IdentificacionAsesor"]),
                    _s(row["CodigoEstado"]),
                    _s(row["CodigoCanal"]),
                    _s(row["CodigoGestion"]),
                    _s(row["TelefonoContacto"]),
                    _s(row["Direccion"]),
                    _s(row["FechaPromesa"]),
                    _s(row["ValorPromesa"]),
                    _s(row["NumeroObligacion"]),
                    _s(row["FechaHoraGestion"]),
                    _s(row["Observaciones"]),
                )
                param_rows.append(params)
                
            except Exception as row_error:
                print(f"❌ Error procesando fila {idx + 1}: {row_error}")
                print(f"Datos de la fila: {row.to_dict()}")
                continue
        
        if not param_rows:
            raise ValueError("No se pudo procesar ninguna fila del Excel")
        insertados, total = insertar_acuerdos_lote(param_rows, chunk_size=100)
        return {"total": total, "insertados_ok": insertados, "dni_list": dni_list}
        
    except Exception as e:
        error_msg = f"Error procesando Excel: {str(e)}"
        print(f"🔥 {error_msg}")
        raise ValueError(error_msg)
    
def get_dni(estado: Optional[str] = "PENDIENTE") -> str:
    dni = obtener_dni(estado)
    if not dni:
        raise ValueError("No hay DNIs disponibles para el estado solicitado")
    return dni

def pop_dni_get(estado_from: Optional[str] = "PENDIENTE",
            estado_to: Optional[str] = "ENVIADO") -> str:
    """
    Obtiene 1 DNI aleatorio y cambia su estado de estado_from → estado_to.
    Lanza ValueError si no hay disponible.
    """
    dni = pop_dni(estado_from, estado_to)
    if not dni:
        raise ValueError("No hay DNIs disponibles para el estado solicitado")
    return dni

def obtener_registro(
    *, solo_activos: bool = True, exige_estado: Optional[str] = "ACTIVO"
) -> Optional[Dict[str, Any]]:
    """
    Orquesta la operación y devuelve el registro actualizado,
    o None si no había pendientes.
    """
    return pop_acuerdo(solo_activos=solo_activos, exige_estado=exige_estado)